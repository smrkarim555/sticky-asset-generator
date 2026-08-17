import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv

topics_30 = [
    (1, "Sticky Note", "Paper Assets"),
    (2, "Kraft Paper", "Paper Assets"),
    (3, "Torn Paper", "Paper Assets"),
    (4, "Gold Pin", "Paper Assets"),
    (5, "Scotch Tape", "Paper Assets"),
    (6, "Paper Clip", "Paper Assets"),
    (7, "Lined Memo", "Paper Assets"),
    (8, "Stage Spotlight", "Light Effects"),
    (9, "Volumetric Fog", "Light Effects"),
    (10, "God Ray", "Light Effects"),
    (11, "Sun Beam", "Light Effects"),
    (12, "Neon Light", "Light Effects"),
    (13, "Laser Beam", "Light Effects"),
    (14, "Powder Burst", "Powder & Sand"),
    (15, "Sand Blast", "Powder & Sand"),
    (16, "Dust Cloud", "Powder & Sand"),
    (17, "Holi Color", "Powder & Sand"),
    (18, "Gold Confetti", "Confetti & FX"),
    (19, "Silver Foil", "Confetti & FX"),
    (20, "Party Ribbon", "Confetti & FX"),
    (21, "Glitter Spray", "Confetti & FX"),
    (22, "Lens Flare", "Confetti & FX"),
    (23, "Speed Lines", "Action & Art"),
    (24, "Anime Burst", "Action & Art"),
    (25, "Brush Stroke", "Action & Art"),
    (26, "Ink Splash", "Action & Art"),
    (27, "Particle Wave", "Action & Art"),
    (28, "Smoke Trail", "Powder & Sand"),
    (29, "Prism Ray", "Light Effects"),
    (30, "Light Shaft", "Light Effects")
]

# Create XLSX
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30 Search Topics"
ws.views.sheetView[0].showGridLines = True

headers = ["SL", "2-Word Google Search Topic", "Asset Category"]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for r_idx, (sl, topic, cat) in enumerate(topics_30, start=2):
    fill = alt_fill if r_idx % 2 == 0 else white_fill
    ws.append([sl, topic, cat])
    for c_idx in range(1, 4):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = fill
        if c_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20

wb.save("PNG_30_Topics_Search.xlsx")

# Create CSV
with open("PNG_30_Topics_Search.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for sl, topic, cat in topics_30:
        writer.writerow([sl, topic, cat])

print("Successfully generated 30 topics Excel and CSV files.")
