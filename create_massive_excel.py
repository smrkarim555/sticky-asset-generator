import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv

categories = {
    "1. Sticky Notes & Paper Cutouts": [
        "Yellow Sticky Note", "Pink Sticky Note", "Mint Sticky Note", "Blue Sticky Note", "Orange Sticky Note",
        "Purple Sticky Note", "Neon Sticky Note", "Pastel Sticky Note", "Square Paper Note", "Torn Edge Paper",
        "Deckled Edge Paper", "Kraft Paper Scrap", "Recycled Brown Paper", "Lined Notebook Paper", "Grid Graph Paper",
        "Dot Grid Paper", "Vintage Parchment", "Burnt Edge Paper", "Folded Paper Scrap", "Crushed Paper Ball",
        "Spiral Memo Sheet", "Index Card Note", "Receipt Paper Cutout", "Cardboard Scrap", "Newsprint Clipping",
        "Paper Tag Label", "Envelope Scrap", "Postcard Cutout", "Bookmark Strip", "Origami Fold Note",
        "Watercolor Paper Scrap", "Handmade Fiber Paper", "Gold Foil Edge Paper", "Black Paper Note", "Translucent Vellum Sheet"
    ],
    "2. Fasteners, Pins & Tapes": [
        "Silver Pushpin", "Gold Pushpin", "Red Thumbtack", "Blue Thumbtack", "Black Pushpin",
        "Wooden Pushpin", "Transparent Pushpin", "Silver Paperclip", "Gold Paperclip", "Black Binder Clip",
        "Copper Paperclip", "Bulldog Metal Clip", "Safety Pin Metal", "Brass Fastener Brad", "Magnet Button Anchor",
        "Scotch Tape Strip", "Washi Tape Floral", "Grid Washi Tape", "Brown Masking Tape", "Clear Cellophane Tape",
        "Duct Tape Strip", "Patterned Craft Tape", "Corner Photo Sticker", "Wax Seal Stamp", "Adhesive Sticker Tag"
    ],
    "3. Stage Lights & Spotlights": [
        "Stage Cone Spotlight", "Diagonal Light Shaft", "Broad Stage Spotlight", "Narrow Laser Spotlight", "Prism Rainbow Spotlight",
        "Cyberpunk Neon Spotlight", "Warm Stage Spotlight", "Cool White Spotlight", "Dual Stage Spotlights", "Triple Spotlight Array",
        "Top Down Spotlight", "Bottom Uplight Beam", "Center Focal Spotlight", "Soft Edge Spotlight", "Hard Edge Beam",
        "Atmospheric Light Cone", "Volumetric Stage Haze", "Stage Dust Particles", "Concert Lighting Shaft", "Theater Stage Beam"
    ],
    "4. Sunbeams, Rays & Flares": [
        "Golden God Ray", "Atmospheric Sunbeam", "Forest Canopy Sunray", "Cloud Light Dispersion", "Lens Flare Sparkle",
        "Optical Star Flare", "Anamorphic Flare Streak", "Sunburst Ray Burst", "Prism Spectrum Ray", "Rainbow Light Flare",
        "Window Light Shadow", "Blinds Light Pattern", "Tree Shadow Light", "Morning Sun Ray", "Sunset Golden Light"
    ],
    "5. Neon, Cyberpunk & Lasers": [
        "Cyan Neon Glow", "Magenta Laser Beam", "Cyberpunk Grid Glow", "Neon Circle Ring", "Neon Frame Border",
        "Holographic Light Ray", "Plasma Energy Wave", "Laser Grid Lines", "Electric Spark Glow", "Fibre Optic Strands",
        "Glowing Portal Ring", "Futuristic Light Bar", "Synthwave Light Grid", "Neon Arrow Icon", "Neon Crosshair Target"
    ],
    "6. Powders, Sand & Particles": [
        "Sand Explosion Blast", "Desert Sand Wave", "Golden Grain Spray", "Cyan Powder Burst", "Magenta Powder Cloud",
        "Holi Color Blast", "White Flour Explosion", "Black Charcoal Dust", "Crushed Spice Powder", "Volcanic Ash Cloud",
        "Dispersed Particle Blast", "Disintegration Effect", "Dust Mote Particles", "Floating Gold Dust", "Cosmic Dust Spray"
    ],
    "7. Smoke, Fog & Clouds": [
        "Dense White Smoke", "Thin Smoke Trail", "Whispy Smoke Swirl", "Colored Smoke Ring", "Cinematic Fog Overlay",
        "Ground Heavy Fog", "Atmospheric Haze Layer", "Cloud Cluster Cutout", "Cumulus Cloud Puff", "Storm Cloud Fragment",
        "Steam Vapor Burst", "Dry Ice Fog Wave", "Exhaust Smoke Wisps", "Explosion Smoke Plume", "Mystical Magic Smoke"
    ],
    "8. Confetti, Ribbons & Party": [
        "Gold Foil Confetti", "Silver Foil Confetti", "Multicolor Paper Confetti", "Confetti Burst Explosion", "Curled Ribbon Streamer",
        "Gold Metallic Ribbon", "Party Popper Spray", "Celebration Star Burst", "Sparkling Glitter Cloud", "Championship Ticker Tape",
        "Pastel Confetti Dots", "Geometric Confetti Squares", "Heart Shaped Confetti", "Circle Paper Confetti", "Floating Party Ribbons"
    ],
    "9. Glitter, Sparkles & Flares": [
        "Gold Glitter Spray", "Silver Sparkle Dust", "Holographic Glitter", "Diamond Sparkle Flare", "Magic Fairy Dust",
        "Bokeh Light Orbs", "Glowing Particle Cluster", "Sparkler Fire Effect", "Golden Star Sparkles", "Cosmic Stardust Cloud"
    ],
    "10. Anime, Speed Lines & Action": [
        "Radial Speed Spikes", "Horizontal Motion Blur", "Manga Impact Burst", "Comic Action Lines", "Focus Zoom Rays",
        "Kinetic Energy Slash", "Sword Slice Effect", "Lightning Bolt Burst", "Shockwave Ring Wave", "Fireball Burst Particle"
    ],
    "11. Ink, Paint & Splatters": [
        "Black Ink Splatter", "Color Paint Drip", "Watercolor Texture Wash", "Dry Brush Stroke", "Calligraphy Ink Blot",
        "Acrylic Paint Splash", "Spray Paint Mist", "Stamp Ink Imprint", "Marbled Ink Swirl", "Graffiti Paint Spray"
    ],
    "12. Frames, Shapes & Badges": [
        "Gold Metallic Frame", "Vintage Ornament Frame", "Grunge Distressed Border", "Geometric Poly Ring", "Badge Emblem Cutout",
        "Ribbon Banner Label", "Sale Tag Cutout", "Discount Starburst", "Neon Sign Frame", "Ripped Paper Frame"
    ],
    "13. Nature, Fire & Elements": [
        "Flame Fire Spark", "Ember Particle Glow", "Water Splash Droplet", "Water Ripple Ring", "Ice Crystal Frost",
        "Lightning Spark Branch", "Falling Green Leaves", "Autumn Dried Leaf", "Flower Petal Burst", "Feather Floating Cutout"
    ]
}

# Flatten items
flat_data = []
item_id = 1
for cat, items in categories.items():
    for item in items:
        flat_data.append({
            "id": item_id,
            "category": cat.split(". ")[1],
            "item": item,
            "search_query": f"{item} transparent PNG 4K",
            "format": "4K PNG (32-bit RGBA)"
        })
        item_id += 1

print(f"Total Items Generated: {len(flat_data)}")

# Create Excel (.xlsx)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Master PNG Asset List"
ws.views.sheetView[0].showGridLines = True

# Title Header
ws.merge_cells('A1:E1')
t_cell = ws['A1']
t_cell.value = f"🚀 MASTER 4K PNG GRAPHIC ASSET & EFFECT DIRECTORY ({len(flat_data)} TOTAL TOPICS)"
t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
t_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# Subtitle Header
ws.merge_cells('A2:E2')
s_cell = ws['A2']
s_cell.value = "📌 Search any 2-3 word topic in Google Images -> Upload Reference Image to Gemini AI Studio -> Get 4K Transparent PNG"
s_cell.font = Font(name="Calibri", size=10, italic=True, color="D9E1F2")
s_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
s_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 24

# Column Headers
headers = ["ID", "Category", "2-3 Word Search Topic", "Recommended Google Search Query", "Output Spec"]
ws.row_dimensions[3].height = 26

h_fill = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid")
h_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col_idx)
    c.value = h
    c.fill = h_fill
    c.font = h_font
    c.alignment = Alignment(horizontal="center", vertical="center")

# Rows
alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for r_idx, row_data in enumerate(flat_data, start=4):
    ws.row_dimensions[r_idx].height = 20
    fill = alt_fill if r_idx % 2 == 0 else white_fill
    
    r_vals = [row_data["id"], row_data["category"], row_data["item"], row_data["search_query"], row_data["format"]]
    for c_idx, val in enumerate(r_vals, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.fill = fill
        cell.font = Font(name="Calibri", size=10)
        if c_idx in [1, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 42
ws.column_dimensions["E"].width = 22

wb.save("PNG_Master_All_Assets_List.xlsx")

# Create CSV
with open("PNG_Master_All_Assets_List.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for r in flat_data:
        writer.writerow([r["id"], r["category"], r["item"], r["search_query"], r["format"]])

print("Successfully created master Excel and CSV files.")
