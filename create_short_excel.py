import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv

short_data = [
    # Category 1: Sticky Notes & Paper
    {"id": 1, "category": "Sticky Notes", "item": "Yellow Sticky Note + Silver Pin", "search_tag": "sticky note, yellow memo, pin, paper cutout", "format": "4K PNG"},
    {"id": 2, "category": "Sticky Notes", "item": "Kraft Paper Torn Bottom + Tape", "search_tag": "kraft paper, torn edge, scotch tape, vintage", "format": "4K PNG"},
    {"id": 3, "category": "Sticky Notes", "item": "Pastel Pink Memo + Gold Pin", "search_tag": "pastel pink, memo note, gold pin, aesthetic", "format": "4K PNG"},
    {"id": 4, "category": "Sticky Notes", "item": "Pastel Mint Green + Cellophane Tape", "search_tag": "mint green, note, clear tape, UI widget", "format": "4K PNG"},
    {"id": 5, "category": "Sticky Notes", "item": "Lined Notebook Scrap + Red Pin", "search_tag": "lined paper, notebook torn, red pushpin", "format": "4K PNG"},
    {"id": 6, "category": "Sticky Notes", "item": "Neon Orange Alert Sticky Note", "search_tag": "neon orange, alert banner, shadow projection", "format": "4K PNG"},

    # Category 2: Light & Spotlights
    {"id": 7, "category": "Light Beams", "item": "45° Diagonal Stage Spotlight Beam", "search_tag": "spotlight, light shaft, volumetric beam, stage light", "format": "4K PNG"},
    {"id": 8, "category": "Light Beams", "item": "Broad Stage Illumination Cone", "search_tag": "stage beam, light haze, concert light, atmosphere", "format": "4K PNG"},
    {"id": 9, "category": "Light Beams", "item": "Warm Golden Hour Sunbeam (God Ray)", "search_tag": "sunbeam, god ray, golden light, nature ray", "format": "4K PNG"},
    {"id": 10, "category": "Light Beams", "item": "Cyberpunk Neon Dual Beam (Cyan/Magenta)", "search_tag": "neon light, cyberpunk ray, laser spotlight, gaming", "format": "4K PNG"},
    {"id": 11, "category": "Light Beams", "item": "Focused Pinpoint Laser Spotlight", "search_tag": "laser spot, narrow beam, jewelry highlight", "format": "4K PNG"},

    # Category 3: Sand & Powder Explosions
    {"id": 12, "category": "Sand & Powder", "item": "Dynamic Sand Blast & Powder Explosion", "search_tag": "sand blast, powder explosion, dust cloud, particle burst", "format": "4K PNG"},
    {"id": 13, "category": "Sand & Powder", "item": "Golden Desert Sand Storm Wave", "search_tag": "desert sand, sand wave, dust storm, particles", "format": "4K PNG"},
    {"id": 14, "category": "Sand & Powder", "item": "Holi Color Powder Cloud Burst", "search_tag": "color powder, holi burst, cyan magenta dust, festival", "format": "4K PNG"},

    # Category 4: Confetti & FX
    {"id": 15, "category": "Confetti & FX", "item": "Metallic Gold & Silver Foil Confetti", "search_tag": "confetti burst, gold foil, silver confetti, celebration", "format": "4K PNG"},
    {"id": 16, "category": "Confetti & FX", "item": "Multi-Color Party Ribbon Streamers", "search_tag": "party ribbons, streamers, rainbow confetti, birthday", "format": "4K PNG"},
    {"id": 17, "category": "Confetti & FX", "item": "Shimmering Glitter Dust Particle Burst", "search_tag": "glitter spray, sparkle dust, lens flare, beauty FX", "format": "4K PNG"},
    {"id": 18, "category": "Confetti & FX", "item": "Manga Anime Radial Speed Spikes", "search_tag": "speed lines, anime burst, manga spikes, action lines", "format": "4K PNG"},
    {"id": 19, "category": "Confetti & FX", "item": "Kinetic Ink Dry-Brush Stroke Cluster", "search_tag": "ink splash, brush stroke, dry brush, typography FX", "format": "4K PNG"},

    # Category 5: AI Custom Assets
    {"id": 20, "category": "AI Custom FX", "item": "Gemini AI Style-Matched 4K Cutouts", "search_tag": "gemini AI asset, style match, custom PNG generator", "format": "4K PNG"}
]

# Write XLSX
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Short PNG Asset List"
ws.views.sheetView[0].showGridLines = True

headers = ["ID", "Category", "PNG Asset / Effect", "Search Keywords / Tags", "Resolution & Format"]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for r_idx, row in enumerate(short_data, start=2):
    fill = alt_fill if r_idx % 2 == 0 else white_fill
    r_vals = [row["id"], row["category"], row["item"], row["search_tag"], row["format"]]
    ws.append(r_vals)
    for c_idx in range(1, 6):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = fill
        if c_idx in [1, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 18

wb.save("PNG_Short_Search_List.xlsx")

# Write CSV
with open("PNG_Short_Search_List.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for r in short_data:
        writer.writerow([r["id"], r["category"], r["item"], r["search_tag"], r["format"]])

print("Successfully created short search list files.")
