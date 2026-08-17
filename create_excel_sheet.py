import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os

# Define dataset
data = [
    {
        "sl": 1,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "Classic Canary Yellow Sticky Note with Silver Pushpin",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "2.5 MB - 4.8 MB",
        "visual_details": "Realistic paper fibers (750k particles), 3D drop shadow, metallic silver thumbtack anchor, smooth rounded corners.",
        "use_case": "UI design mockups, noticeboards, website hero sections, task lists, reminder banners.",
        "preset_name": "yellow_sticky_note_silver_pin.png"
    },
    {
        "sl": 2,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "Torn Edge Vintage Kraft Paper Note with Translucent Tape",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "3.2 MB - 6.5 MB",
        "visual_details": "Deckled/torn bottom edge texture, recycled paper grain, matte adhesive tape strip fixed at 15-degree angle.",
        "use_case": "Vintage aesthetic posters, scrapbooks, organic brand designs, moodboards.",
        "preset_name": "kraft_paper_torn_bottom_tape.png"
    },
    {
        "sl": 3,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "Soft Pastel Pink Pinned Memo with Gold Pushpin",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "2.4 MB - 5.0 MB",
        "visual_details": "Soft pastel shade, subtle corner curl shadow, high-specular gold metallic pin, ultra-clean edges.",
        "use_case": "Social media graphics, aesthetic Instagram stories, lifestyle & fashion branding.",
        "preset_name": "pastel_pink_gold_pin_memo.png"
    },
    {
        "sl": 4,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "Pastel Mint Green Note with Transparent Scotch Tape",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "2.6 MB - 5.2 MB",
        "visual_details": "Mint green paper surface, realistic cellophane tape with micro bubbles and edge reflection.",
        "use_case": "Product feature callouts, educational apps, modern web dashboard widgets.",
        "preset_name": "mint_green_tape_sticky_note.png"
    },
    {
        "sl": 5,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "Lined Notebook Scrap Paper with Red Pushpin Anchor",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "3.0 MB - 5.8 MB",
        "visual_details": "Ruled horizontal lines, torn notebook spiral perforations, vivid red glossy pushpin with drop shadow.",
        "use_case": "Student portals, note-taking apps, editorial graphics, tutorial graphics.",
        "preset_name": "lined_notebook_scrap_red_pin.png"
    },
    {
        "sl": 6,
        "category": "Pinned Sticky Notes & Paper Cutouts",
        "title": "High-Contrast Neon Orange Alert Note",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "2.8 MB - 5.5 MB",
        "visual_details": "Vivid vibrant orange neon hue, deep ambient occlusion shadow projection for high pop-out contrast.",
        "use_case": "Urgent notifications, discount promo badges, flash sale callouts.",
        "preset_name": "neon_orange_alert_sticky_note.png"
    },
    {
        "sl": 7,
        "category": "Volumetric Stage Spotlight Studio",
        "title": "Classic 45-Degree Stage Spotlight Beam & Cone Shaft",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.5 MB - 8.5 MB",
        "visual_details": "Volumetric light dust particles, dramatic cone ray, soft radial falloff gradient, true transparent background blend.",
        "use_case": "Product stage presentations, award ceremony banners, hero product showcases.",
        "preset_name": "stage_spotlight_45deg_cone.png"
    },
    {
        "sl": 8,
        "category": "Volumetric Stage Spotlight Studio",
        "title": "Broad Dramatic Illumination Light Shaft",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "5.0 MB - 9.2 MB",
        "visual_details": "Wide aperture beam, realistic haze particles, multi-layered volumetric fog texture, smooth alpha opacity transition.",
        "use_case": "Concert posters, theatrical lighting effects, cinematic background overlays.",
        "preset_name": "broad_stage_illumination_shaft.png"
    },
    {
        "sl": 9,
        "category": "Volumetric Stage Spotlight Studio",
        "title": "Warm Golden Hour Atmospheric God Ray / Sunbeam",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.8 MB - 8.9 MB",
        "visual_details": "5600K warm golden temperature light, glowing micro-dust motes, natural sunlight beam dispersion.",
        "use_case": "Nature & outdoor product ads, spiritual or dreamy visual compositions.",
        "preset_name": "golden_hour_sunbeam_god_ray.png"
    },
    {
        "sl": 10,
        "category": "Volumetric Stage Spotlight Studio",
        "title": "Cyberpunk Neon Blue & Magenta Dual Spotlight Beam",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "5.2 MB - 9.8 MB",
        "visual_details": "Futuristic dual chromatic beam, high-contrast cyan/magenta laser particulate spray, glowing volumetric core.",
        "use_case": "Gaming thumbnail overlays, esports graphics, synthwave/cyberpunk art.",
        "preset_name": "cyberpunk_neon_dual_spotlight.png"
    },
    {
        "sl": 11,
        "category": "Volumetric Stage Spotlight Studio",
        "title": "Narrow Focused Pinpoint Spotlight Laser Beam",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "3.8 MB - 7.2 MB",
        "visual_details": "Tight concentrated beam core, sharp edge boundary, high intense focal brightness, floating dust motes.",
        "use_case": "High-end luxury jewelry product highlight, mystery element spotlighting.",
        "preset_name": "narrow_pinpoint_laser_spotlight.png"
    },
    {
        "sl": 12,
        "category": "Sand, Powder & Dust Explosion Studio",
        "title": "Dynamic Sand Blast & High-Velocity Powder Explosion",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "5.5 MB - 9.9 MB",
        "visual_details": "High-density micro-particle burst, organic liquid grain texture, directional kinetic dispersion clouds.",
        "use_case": "Action movie posters, sports apparel advertisements, energetic motion graphics.",
        "preset_name": "sand_powder_explosion_blast.png"
    },
    {
        "sl": 13,
        "category": "Sand, Powder & Dust Explosion Studio",
        "title": "Golden Desert Sand Storm Wave & Particles",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.9 MB - 8.7 MB",
        "visual_details": "Fluid velocity sand waves, golden ochre grain spray, layered depth-blur sand particles.",
        "use_case": "Adventure travel banners, vehicle promotions, desert-themed branding.",
        "preset_name": "golden_sand_storm_wave.png"
    },
    {
        "sl": 14,
        "category": "Sand, Powder & Dust Explosion Studio",
        "title": "Vibrant Holi Color Powder Cloud Burst (Cyan & Pink)",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "6.0 MB - 9.9 MB",
        "visual_details": "Two-tone pigment powder explosion, fine powder mist, high macro texture resolution.",
        "use_case": "Festival graphic design, music album art, colorful brand launch posters.",
        "preset_name": "holi_color_powder_cloud_burst.png"
    },
    {
        "sl": 15,
        "category": "Confetti & Celebration Burst Studio",
        "title": "Metallic Gold & Silver Foil Confetti Burst",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.2 MB - 8.1 MB",
        "visual_details": "3D metallic foil rectangles & squares, specular metallic reflections, depth-of-field blur on foreground pieces.",
        "use_case": "New Year celebrations, anniversary announcements, winner podium graphics, promotional offers.",
        "preset_name": "metallic_gold_silver_confetti.png"
    },
    {
        "sl": 16,
        "category": "Confetti & Celebration Burst Studio",
        "title": "Multi-Color Ribbon Streamer & Party Dot Spray",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.0 MB - 7.8 MB",
        "visual_details": "Curled ribbon streamers, rainbow confetti circles, floating dynamic motion layout.",
        "use_case": "Birthday party invitations, festive banners, e-commerce sale announcements.",
        "preset_name": "multicolor_ribbon_streamers.png"
    },
    {
        "sl": 17,
        "category": "Confetti & Celebration Burst Studio",
        "title": "Shimmering Glitter Dust Particle Explosion",
        "resolution": "3840 x 2160 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "5.1 MB - 9.5 MB",
        "visual_details": "Micro glitter stars, glowing lens flares, sparkling particle dust spray with transparent alpha channel.",
        "use_case": "Cosmetics & beauty ad graphics, luxury award banners, holiday greeting cards.",
        "preset_name": "shimmering_glitter_dust_explosion.png"
    },
    {
        "sl": 18,
        "category": "Confetti & Celebration Burst Studio",
        "title": "Manga / Comic Action Kinetic Radial Speed Spikes",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "3.5 MB - 6.8 MB",
        "visual_details": "Radial zoom speed lines, bold anime/manga impact burst, sharp line vectors with gradient shading.",
        "use_case": "Manga comic illustrations, anime thumbnails, high-energy product launch highlights.",
        "preset_name": "manga_radial_speed_spikes.png"
    },
    {
        "sl": 19,
        "category": "Confetti & Celebration Burst Studio",
        "title": "Kinetic Ink Dry-Brush Stroke Cluster",
        "resolution": "3840 x 3840 (4K UHD)",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "4.1 MB - 7.9 MB",
        "visual_details": "180+ kinetic brush strokes, 300+ bristle lines, organic ink texture, dry-brush edge distress.",
        "use_case": "Calligraphy backgrounds, typography emphasis, martial arts or fitness graphic posters.",
        "preset_name": "kinetic_ink_brush_cluster.png"
    },
    {
        "sl": 20,
        "category": "Gemini AI Custom Asset Generator",
        "title": "Gemini AI Style-Matched 4K Cutout Asset Set (4 Variations)",
        "resolution": "3840 x 2160 / 3840 x 3840",
        "transparency": "True 32-bit RGBA Alpha",
        "file_size": "2.0 MB - 10.0 MB",
        "visual_details": "AI visual style analysis of reference image -> Automatic generation of 4 matching 4K transparent PNGs.",
        "use_case": "Custom client branding, bespoke graphic asset creation, automated design expansion.",
        "preset_name": "gemini_ai_custom_style_variations.png"
    }
]

# 1. Create Excel Workbook (.xlsx)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PNG Asset Master List"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Title Header Block
ws.merge_cells('A1:I1')
title_cell = ws['A1']
title_cell.value = "🎨 4K PNG ASSET GENERATOR - BEST OUTPUT SPECIFICATION MASTER LIST"
title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Subtitle / Rule Banner Block
ws.merge_cells('A2:I2')
sub_cell = ws['A2']
sub_cell.value = "📌 Rule Specs: 4K UHD Resolution | True 32-bit RGBA Alpha Transparency | Guaranteed File Size: 2.0 MB - 10.0 MB | Lossless Cutouts"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color="D9E1F2")
sub_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 24

# Column Headers
headers = [
    "SL No", 
    "Asset Category", 
    "PNG Asset Title & Description", 
    "Resolution / Specs", 
    "Alpha Transparency", 
    "Target File Size Range", 
    "Visual Details & Features", 
    "Ideal Use Case / Application", 
    "Output File Preset Name"
]

header_fill = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

ws.row_dimensions[3].height = 28
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Data Rows
alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row_idx, item in enumerate(data, start=4):
    ws.row_dimensions[row_idx].height = 36
    row_fill = alt_fill if row_idx % 2 == 0 else white_fill
    
    r_vals = [
        item["sl"],
        item["category"],
        item["title"],
        item["resolution"],
        item["transparency"],
        item["file_size"],
        item["visual_details"],
        item["use_case"],
        item["preset_name"]
    ]
    
    for col_idx, val in enumerate(r_vals, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = val
        cell.font = Font(name="Calibri", size=10)
        cell.fill = row_fill
        cell.border = thin_border
        
        if col_idx in [1, 4, 5, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Set Column Widths
col_widths = {
    "A": 8,   # SL No
    "B": 28,  # Category
    "C": 35,  # Title
    "D": 24,  # Resolution
    "E": 22,  # Transparency
    "F": 20,  # File Size
    "G": 45,  # Visual Details
    "H": 38,  # Use Case
    "I": 32   # Output File Name
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save Excel File
xlsx_path = "PNG_Best_Asset_Outputs_List.xlsx"
wb.save(xlsx_path)
print(f"Successfully generated Excel file: {xlsx_path}")

# 2. Save CSV File (UTF-8 with BOM for Excel compatibility)
csv_path = "PNG_Best_Asset_Outputs_List.csv"
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for item in data:
        writer.writerow([
            item["sl"],
            item["category"],
            item["title"],
            item["resolution"],
            item["transparency"],
            item["file_size"],
            item["visual_details"],
            item["use_case"],
            item["preset_name"]
        ])
print(f"Successfully generated CSV file: {csv_path}")
